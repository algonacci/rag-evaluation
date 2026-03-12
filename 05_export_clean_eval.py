import argparse
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


INPUT_CSV = "04_evaluated.csv"
OUTPUT_XLSX = "05_evaluated_clean.xlsx"
USED_COLUMNS = [
    "question",
    "context",
    "ground_truth_answer",
    "generated_answer",
    "context_precision",
    "context_recall",
    "context_entity_recall",
    "faithfulness",
    "answer_relevancy",
]
DISPLAY_NAMES = {
    "question": "question",
    "context": "context",
    "ground_truth_answer": "ground_truth_answer",
    "generated_answer": "generated_answer",
    "context_precision": "context_precision",
    "context_recall": "context_recall",
    "context_entity_recall": "context_entity_recall",
    "faithfulness": "faithfulness",
    "answer_relevancy": "answer_relevancy",
    "evaluation_status": "status",
}
WIDTH_LIMITS = {
    "A": 42,
    "B": 70,
    "C": 60,
    "D": 60,
    "E": 18,
    "F": 18,
    "G": 16,
    "H": 18,
    "I": 14,
}


def parse_metric_value(value):
    if pd.isna(value):
        return None

    text = str(value).strip()
    if not text or text.lower() == "none":
        return None

    match = re.search(r"MetricResult\(value=([-+]?\d*\.?\d+(?:[eE][-+]?\d+)?)\)", text)
    if match:
        return float(match.group(1))

    try:
        return float(text)
    except ValueError:
        return None


def adjust_excel_layout(path):
    workbook = load_workbook(path)
    worksheet = workbook["evaluation"]
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter in {"A", "B", "C", "D"}:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")

    for column_cells in worksheet.columns:
        column_letter = column_cells[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, WIDTH_LIMITS.get(column_letter, 20))

    if "summary" in workbook.sheetnames:
        summary_sheet = workbook["summary"]
        summary_sheet.freeze_panes = "A2"
        for cell in summary_sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in summary_sheet.columns:
            column_letter = column_cells[0].column_letter
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            summary_sheet.column_dimensions[column_letter].width = min(max_length + 2, 32)

    workbook.save(path)


parser = argparse.ArgumentParser(description="Export a clean evaluation Excel file.")
parser.add_argument("--input", default=INPUT_CSV, help="Input evaluation CSV.")
parser.add_argument("--output", default=OUTPUT_XLSX, help="Output Excel path.")
args = parser.parse_args()

df = pd.read_csv(args.input)

# Normalize legacy column names from older evaluation outputs.
if "context_entity_recall" not in df.columns and "context_entities_recall" in df.columns:
    df["context_entity_recall"] = df["context_entities_recall"]
if "answer_relevancy" not in df.columns and "response_relevancy" in df.columns:
    df["answer_relevancy"] = df["response_relevancy"]

for column in [
    "context_precision",
    "context_recall",
    "context_entity_recall",
    "faithfulness",
    "answer_relevancy",
]:
    df[column] = df[column].apply(parse_metric_value)

df = df[USED_COLUMNS].copy()
metric_columns = [
    "context_precision",
    "context_recall",
    "context_entity_recall",
    "faithfulness",
    "answer_relevancy",
]
df = df[df[metric_columns].notna().any(axis=1)].copy()
df["evaluation_status"] = df[metric_columns].notna().all(axis=1).map(
    lambda is_complete: "complete" if is_complete else "partial"
)

complete_count = int((df["evaluation_status"] == "complete").sum())
partial_count = int((df["evaluation_status"] == "partial").sum())

df = df.rename(columns=DISPLAY_NAMES)

summary_rows = [
    {"metric": "rows_exported", "value": len(df)},
    {"metric": "rows_complete_5_metrics", "value": complete_count},
    {"metric": "rows_partial_5_metrics", "value": partial_count},
]
for column in [
    "context_precision",
    "context_recall",
    "context_entity_recall",
    "faithfulness",
    "answer_relevancy",
]:
    valid = df[column].dropna()
    summary_rows.append(
        {
            "metric": f"avg_{column}",
            "value": round(valid.mean(), 4) if not valid.empty else None,
        }
    )

summary_rows.append(
    {
        "metric": "notes",
        "value": "noise_sensitivity intentionally excluded from the final clean report due to repeated evaluator timeouts",
    }
)
summary_df = pd.DataFrame(summary_rows)

with pd.ExcelWriter(args.output, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="evaluation")
    summary_df.to_excel(writer, index=False, sheet_name="summary")

adjust_excel_layout(args.output)

print(f"Clean Excel exported to {args.output}")
print(f"Rows exported: {len(df)}")
